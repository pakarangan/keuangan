from fastapi import FastAPI, HTTPException, Depends, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, validator
from typing import List, Optional, Dict
from datetime import datetime, date
from decimal import Decimal
from bson import ObjectId
from passlib.context import CryptContext
from jose import JWTError, jwt
import os
import logging
from pathlib import Path
from dotenv import load_dotenv
import uuid

# Load environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
import hashlib
security = HTTPBearer()
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-here")
ALGORITHM = "HS256"

# Create FastAPI app
app = FastAPI(
    title="Financial Records API",
    description="API untuk aplikasi pencatatan keuangan",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

# Data Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    hashed_password: str
    company_name: str
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    username: str
    email: str
    password: str
    company_name: str

class UserLogin(BaseModel):
    username: str
    password: str

class Account(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    name: str
    code: Optional[str] = None
    category: str  # Aset, Utang, Pendapatan, Modal, Biaya
    balance: Decimal = Decimal('0.00')
    created_at: datetime = Field(default_factory=datetime.utcnow)

class AccountCreate(BaseModel):
    name: str
    code: Optional[str] = None
    category: str
    
    @validator('category')
    def validate_category(cls, v):
        valid_categories = ['Aset', 'Utang', 'Pendapatan', 'Modal', 'Biaya']
        if v not in valid_categories:
            raise ValueError(f'Category must be one of: {valid_categories}')
        return v

class Transaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    date: date
    description: str
    amount: Decimal
    account_id: str
    account_name: str
    receipt_image: Optional[str] = None  # Base64 encoded image
    created_at: datetime = Field(default_factory=datetime.utcnow)

class TransactionCreate(BaseModel):
    date: date
    description: str
    amount: Decimal
    account_id: str
    receipt_image: Optional[str] = None

class FinancialSummary(BaseModel):
    total_aset: Decimal
    total_utang: Decimal
    total_pendapatan: Decimal
    total_modal: Decimal
    total_biaya: Decimal
    net_income: Decimal

# Helper functions
def verify_password(plain_password, hashed_password):
    # Simple hash verification for MVP
    return hashlib.sha256((plain_password + SECRET_KEY).encode()).hexdigest() == hashed_password

def get_password_hash(password):
    # Simple hash for MVP
    return hashlib.sha256((password + SECRET_KEY).encode()).hexdigest()

def create_access_token(data: dict):
    to_encode = data.copy()
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"username": username})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# API Routes
@app.get("/api/health")
async def health_check():
    return {"status": "healthy", "message": "Financial Records API is running"}

@app.post("/api/auth/register")
async def register_user(user_data: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({
        "$or": [
            {"username": user_data.username},
            {"email": user_data.email}
        ]
    })
    
    if existing_user:
        raise HTTPException(status_code=400, detail="Username or email already exists")
    
    # Create user
    hashed_password = get_password_hash(user_data.password)
    user_dict = user_data.dict()
    user_dict.pop('password')
    user_dict['hashed_password'] = hashed_password
    user_dict['id'] = str(uuid.uuid4())
    user_dict['created_at'] = datetime.utcnow()
    
    result = await db.users.insert_one(user_dict)
    
    # Create default accounts
    default_accounts = [
        {"name": "Kas", "category": "Aset", "code": "1001"},
        {"name": "Bank", "category": "Aset", "code": "1002"},
        {"name": "Hutang Usaha", "category": "Utang", "code": "2001"},
        {"name": "Penjualan", "category": "Pendapatan", "code": "4001"},
        {"name": "Modal Pribadi", "category": "Modal", "code": "3001"},
        {"name": "Biaya Operasional", "category": "Biaya", "code": "5001"}
    ]
    
    for acc in default_accounts:
        account_dict = {
            "id": str(uuid.uuid4()),
            "user_id": user_dict['id'],
            "name": acc["name"],
            "code": acc["code"],
            "category": acc["category"],
            "balance": Decimal('0.00'),
            "created_at": datetime.utcnow()
        }
        await db.accounts.insert_one({
            **account_dict,
            "balance": float(account_dict["balance"])
        })
    
    return {"message": "User registered successfully", "user_id": user_dict['id']}

@app.post("/api/auth/login")
async def login_user(login_data: UserLogin):
    user = await db.users.find_one({"username": login_data.username})
    
    if not user or not verify_password(login_data.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    access_token = create_access_token(data={"sub": user["username"]})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user["id"],
            "username": user["username"],
            "email": user["email"],
            "company_name": user["company_name"]
        }
    }

@app.get("/api/accounts", response_model=List[Account])
async def get_user_accounts(current_user: dict = Depends(get_current_user)):
    accounts = await db.accounts.find({"user_id": current_user["id"]}).to_list(None)
    return [
        Account(
            id=acc["id"],
            user_id=acc["user_id"],
            name=acc["name"],
            code=acc.get("code"),
            category=acc["category"],
            balance=Decimal(str(acc.get("balance", 0))),
            created_at=acc["created_at"]
        )
        for acc in accounts
    ]

@app.post("/api/accounts", response_model=Account)
async def create_account(account_data: AccountCreate, current_user: dict = Depends(get_current_user)):
    account_dict = account_data.dict()
    account_dict.update({
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "balance": 0.00,
        "created_at": datetime.utcnow()
    })
    
    result = await db.accounts.insert_one(account_dict)
    
    return Account(
        id=account_dict["id"],
        user_id=account_dict["user_id"],
        name=account_dict["name"],
        code=account_dict.get("code"),
        category=account_dict["category"],
        balance=Decimal('0.00'),
        created_at=account_dict["created_at"]
    )

@app.get("/api/transactions", response_model=List[Transaction])
async def get_user_transactions(
    limit: int = Query(50, le=100),
    offset: int = Query(0, ge=0),
    account_id: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["id"]}
    if account_id:
        query["account_id"] = account_id
    
    transactions = await db.transactions.find(query).sort("date", -1).skip(offset).limit(limit).to_list(None)
    
    return [
        Transaction(
            id=txn["id"],
            user_id=txn["user_id"],
            date=txn["date"],
            description=txn["description"],
            amount=Decimal(str(txn["amount"])),
            account_id=txn["account_id"],
            account_name=txn["account_name"],
            receipt_image=txn.get("receipt_image"),
            created_at=txn["created_at"]
        )
        for txn in transactions
    ]

@app.post("/api/transactions", response_model=Transaction)
async def create_transaction(transaction_data: TransactionCreate, current_user: dict = Depends(get_current_user)):
    # Get account info
    account = await db.accounts.find_one({"id": transaction_data.account_id, "user_id": current_user["id"]})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Create transaction
    transaction_dict = transaction_data.dict()
    transaction_dict.update({
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "account_name": account["name"],
        "amount": float(transaction_data.amount),
        "created_at": datetime.utcnow()
    })
    
    result = await db.transactions.insert_one(transaction_dict)
    
    # Update account balance based on category
    amount_change = float(transaction_data.amount)
    if account["category"] in ["Utang", "Pendapatan", "Modal"]:
        # Credit accounts - increase with positive amounts
        new_balance = float(account.get("balance", 0)) + amount_change
    else:
        # Debit accounts (Aset, Biaya) - increase with positive amounts
        new_balance = float(account.get("balance", 0)) + amount_change
    
    await db.accounts.update_one(
        {"id": transaction_data.account_id},
        {"$set": {"balance": new_balance}}
    )
    
    return Transaction(
        id=transaction_dict["id"],
        user_id=transaction_dict["user_id"],
        date=transaction_dict["date"],
        description=transaction_dict["description"],
        amount=Decimal(str(transaction_dict["amount"])),
        account_id=transaction_dict["account_id"],
        account_name=transaction_dict["account_name"],
        receipt_image=transaction_dict.get("receipt_image"),
        created_at=transaction_dict["created_at"]
    )

@app.get("/api/financial-summary", response_model=FinancialSummary)
async def get_financial_summary(current_user: dict = Depends(get_current_user)):
    accounts = await db.accounts.find({"user_id": current_user["id"]}).to_list(None)
    
    summary = {
        "total_aset": Decimal('0'),
        "total_utang": Decimal('0'),
        "total_pendapatan": Decimal('0'),
        "total_modal": Decimal('0'),
        "total_biaya": Decimal('0')
    }
    
    for account in accounts:
        balance = Decimal(str(account.get("balance", 0)))
        category = account["category"]
        
        if category == "Aset":
            summary["total_aset"] += balance
        elif category == "Utang":
            summary["total_utang"] += balance
        elif category == "Pendapatan":
            summary["total_pendapatan"] += balance
        elif category == "Modal":
            summary["total_modal"] += balance
        elif category == "Biaya":
            summary["total_biaya"] += balance
    
    # Calculate net income (Pendapatan - Biaya)
    net_income = summary["total_pendapatan"] - summary["total_biaya"]
    summary["net_income"] = net_income
    
    return FinancialSummary(**summary)

@app.delete("/api/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, current_user: dict = Depends(get_current_user)):
    transaction = await db.transactions.find_one({"id": transaction_id, "user_id": current_user["id"]})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Update account balance (reverse the transaction)
    account = await db.accounts.find_one({"id": transaction["account_id"]})
    if account:
        amount_change = -float(transaction["amount"])  # Reverse the amount
        if account["category"] in ["Utang", "Pendapatan", "Modal"]:
            new_balance = float(account.get("balance", 0)) + amount_change
        else:
            new_balance = float(account.get("balance", 0)) + amount_change
        
        await db.accounts.update_one(
            {"id": transaction["account_id"]},
            {"$set": {"balance": new_balance}}
        )
    
    # Delete transaction
    await db.transactions.delete_one({"id": transaction_id})
    
    return {"message": "Transaction deleted successfully"}

# Report Generation Endpoints
@app.get("/api/reports/profit-loss/pdf")
async def generate_profit_loss_pdf(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user)
):
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    # Get transactions in date range
    start_datetime = datetime.strptime(start_date, "%Y-%m-%d")
    end_datetime = datetime.strptime(end_date, "%Y-%m-%d")
    
    accounts = await db.accounts.find({"user_id": current_user["id"]}).to_list(None)
    
    # Calculate totals by category
    totals = {
        "Pendapatan": 0,
        "Biaya": 0,
        "Aset": 0,
        "Utang": 0,
        "Modal": 0
    }
    
    for account in accounts:
        balance = float(account.get("balance", 0))
        category = account["category"]
        if category in totals:
            totals[category] += balance
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph(f"<b>{current_user['company_name']}</b><br/>LAPORAN LABA RUGI<br/>Periode: {start_date} s/d {end_date}", 
                     styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Revenue section
    revenue_data = [
        ['PENDAPATAN', ''],
        ['Total Pendapatan', f"Rp {totals['Pendapatan']:,.2f}"],
        ['', ''],
        ['BIAYA OPERASIONAL', ''],
        ['Total Biaya', f"Rp {totals['Biaya']:,.2f}"],
        ['', ''],
        ['LABA BERSIH', f"Rp {totals['Pendapatan'] - totals['Biaya']:,.2f}"]
    ]
    
    table = Table(revenue_data, colWidths=[4*inch, 2*inch])
    table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LINEBELOW', (0, 1), (-1, 1), 1, colors.black),
        ('LINEBELOW', (0, 4), (-1, 4), 1, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
    ]))
    
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer, 
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=profit_loss_{start_date}_{end_date}.pdf"}
    )

@app.get("/api/reports/profit-loss/excel")  
async def generate_profit_loss_excel(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user)
):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    # Get data
    accounts = await db.accounts.find({"user_id": current_user["id"]}).to_list(None)
    
    totals = {
        "Pendapatan": 0,
        "Biaya": 0,
        "Aset": 0,
        "Utang": 0,
        "Modal": 0
    }
    
    for account in accounts:
        balance = float(account.get("balance", 0))
        category = account["category"]
        if category in totals:
            totals[category] += balance
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    
    # Headers
    ws['A1'] = current_user['company_name']
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = "LAPORAN LABA RUGI"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = f"Periode: {start_date} s/d {end_date}"
    
    # Data
    row = 5
    ws[f'A{row}'] = "PENDAPATAN"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Total Pendapatan"
    ws[f'B{row}'] = totals['Pendapatan']
    ws[f'B{row}'].number_format = '"Rp "#,##0.00'
    
    row += 2
    ws[f'A{row}'] = "BIAYA OPERASIONAL"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Total Biaya"
    ws[f'B{row}'] = totals['Biaya']
    ws[f'B{row}'].number_format = '"Rp "#,##0.00'
    
    row += 2
    ws[f'A{row}'] = "LABA BERSIH"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = totals['Pendapatan'] - totals['Biaya']
    ws[f'B{row}'].number_format = '"Rp "#,##0.00'
    ws[f'B{row}'].font = Font(bold=True)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=profit_loss_{start_date}_{end_date}.xlsx"}
    )

@app.get("/api/reports/balance-sheet/pdf")
async def generate_balance_sheet_pdf(
    report_date: str = Query(..., description="Report date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user)
):
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    accounts = await db.accounts.find({"user_id": current_user["id"]}).to_list(None)
    
    # Group accounts by category
    categories = {
        "Aset": [],
        "Utang": [],
        "Modal": []
    }
    
    for account in accounts:
        category = account["category"]
        if category in categories:
            categories[category].append(account)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph(f"<b>{current_user['company_name']}</b><br/>NERACA<br/>Per {report_date}", 
                     styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Assets section
    asset_data = [['ASET', '']]
    total_assets = 0
    for account in categories['Aset']:
        balance = float(account.get("balance", 0))
        asset_data.append([account['name'], f"Rp {balance:,.2f}"])
        total_assets += balance
    asset_data.append(['TOTAL ASET', f"Rp {total_assets:,.2f}"])
    
    # Liabilities section  
    asset_data.append(['', ''])
    asset_data.append(['KEWAJIBAN & MODAL', ''])
    
    total_liabilities = 0
    for account in categories['Utang']:
        balance = float(account.get("balance", 0))
        asset_data.append([account['name'], f"Rp {balance:,.2f}"])
        total_liabilities += balance
    
    total_equity = 0
    for account in categories['Modal']:
        balance = float(account.get("balance", 0))
        asset_data.append([account['name'], f"Rp {balance:,.2f}"])
        total_equity += balance
    
    asset_data.append(['TOTAL KEWAJIBAN & MODAL', f"Rp {total_liabilities + total_equity:,.2f}"])
    
    table = Table(asset_data, colWidths=[4*inch, 2*inch])
    table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename=balance_sheet_{report_date}.pdf"}
    )

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)